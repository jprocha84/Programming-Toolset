# -*- coding: utf-8 -*-
"""
Created on Sat Oct 03 20:38:36 2015

@author: TZ72DL
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd

wb = Workbook()
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")


wb = load_workbook('C:\\company\\Research\\company\\Excel\\inputFiles\\ProvisionRequestList (5).xlsx') 

book = xlrd.open_workbook('C:\\Users\\tz72dl\\Downloads\\ProvisionRequestList (5).xlsx')
sheet = book.sheet_by_index(0)
sheet.nrows
sheet.ncols
sheet.cell_value(3,5)



from mmap import mmap,ACCESS_READ
from xlrd import open_workbook
print open_workbook('C:\\company\\Research\\company\\Excel\\inputFiles\\ProvisionRequestList.xls')
with open('C:\\company\\Research\\company\\Excel\\inputFiles\\ProvisionRequestList.xls','rb') as f:
 print open_workbook(
 file_contents=mmap(f.fileno(),0,access=ACCESS_READ)
 )
aString = open('C:\\company\\Research\\company\\Excel\\inputFiles\\ProvisionRequestList.xls','rb').read()
print open_workbook(file_contents=aString)



from xlrd import open_workbook
wb = open_workbook('C:\\company\\Research\\company\\Excel\\inputFiles\\ProvisionRequestList.xls')
for s in wb.sheets():
    print 'Sheet:',s.name
    for row in range(s.nrows):
        values = []
        for col in range(s.ncols):
            values.append(s.cell(row,col).value)
        print ','.join(values)
    print
